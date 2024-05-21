import openpyxl as xl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import absolute_coordinate
import random

#globals 
myPath = "C:\\Users\\Jeffrey\\Python_projects\\ms_office_tools\\"
myFile = "dv_kentucky_derby.xlsx"

#Load Worksheets
wb = xl.load_workbook(myPath + myFile)
ws = wb["Main Page"]
dvSheet = wb["Dropdowns"]
#print(dvSheet.title)

#Add Data Validation
dv_names = DataValidation(type="list",formula1="{0}!$B$2:$B$5".format(quote_sheetname(dvSheet.title)))
ws.add_data_validation(dv_names)
dv_names.add('B2:B10')

dv_horses = DataValidation(type="list",formula1="{0}!$D$2:$D$23".format(quote_sheetname(dvSheet.title)))
ws.add_data_validation(dv_horses)
dv_horses.add('D2:D10')

#find first empty row
def first_empty(sheet):
    first_empty_row = len(list(sheet.rows))
    return first_empty_row + 1

# add a row of data 
get_row = str(first_empty(ws))
ws["B"+ get_row].value = dvSheet["B" + str(random.randint(2,5))].value
ws["C"+ get_row].value = str(random.randint(2,25))
ws["D"+ get_row].value = dvSheet["D" + str(random.randint(2,23))].value
ws["E"+ get_row].value = dvSheet["E" + str(random.randint(2,4))].value

#Save worksheet
wb.save(myPath + myFile)
